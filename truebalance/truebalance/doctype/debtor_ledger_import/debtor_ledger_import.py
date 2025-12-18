from __future__ import annotations
import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils import now_datetime, flt, getdate
import hashlib
import csv
import io
from datetime import datetime, date

try:
    from openpyxl import load_workbook
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


class DebtorLedgerImport(Document):
    pass  # Leave hooks empty; we use buttons


# -----------------------
# Whitelisted server methods
# -----------------------

@frappe.whitelist()
def parse_file_and_preview(docname):
    doc = frappe.get_doc("Debtor Ledger Import", docname)
    rows, parse_log = _parse_file(doc.file_to_upload)
    _build_preview(doc, rows, parse_log)
    frappe.db.commit()
    return {"status": "success", "parsed_rows": len(doc.debtor_import_preview), "log_count": len(doc.import_log)}


@frappe.whitelist()
def start_import(docname):
    doc = frappe.get_doc("Debtor Ledger Import", docname)
    created, skipped = _create_statement_entries(doc)
    doc.import_end_time = now_datetime()
    doc.import_status = "Completed"
    doc.append("import_log", {
        "log_type": "Info",
        "message": _("Created {0} statement entries, skipped {1} duplicates").format(created, skipped),
        "row_index": 0
    })
    doc.save()
    frappe.db.commit()
    return {"status": "success", "created": created, "skipped": skipped}


# -----------------------
# Helpers
# -----------------------

def _hash_row(company, customer_reference, statement_date, credit, debit):
    base = "|".join([
        company or "",
        customer_reference or "",
        statement_date.isoformat() if statement_date else "",
        str(credit or 0),
        str(debit or 0),
    ])
    return hashlib.md5(base.encode("utf-8")).hexdigest()


def _normalize_date(date_value):
    """Convert various date formats to date object"""
    if date_value is None or date_value == "":
        return None
    
    # Already a date object
    if isinstance(date_value, date):
        return date_value
    
    # datetime object - extract date
    if isinstance(date_value, datetime):
        return date_value.date()
    
    # String - try to parse
    if isinstance(date_value, str):
        date_value = date_value.strip()
        if not date_value:
            return None
        try:
            return getdate(date_value)
        except:
            return None
    
    # Number (Excel serial date)
    if isinstance(date_value, (int, float)):
        try:
            # Excel epoch starts at 1900-01-01
            from datetime import timedelta
            excel_epoch = datetime(1899, 12, 30)
            return (excel_epoch + timedelta(days=float(date_value))).date()
        except:
            return None
    
    return None


def _get_header_index(headers, *possible_names):
    """Find the index of a header by checking multiple possible names (case-insensitive)"""
    headers_lower = [h.lower().strip() if h else "" for h in headers]
    for name in possible_names:
        name_lower = name.lower().strip()
        if name_lower in headers_lower:
            return headers_lower.index(name_lower)
    return None


def _parse_file(file_url):
    rows = []
    parse_log = []

    if not file_url:
        parse_log.append({"log_type": "Error", "message": "No file uploaded", "row_index": 0})
        return rows, parse_log

    content = None
    try:
        f = frappe.get_doc("File", {"file_url": file_url})
        if hasattr(f, "get_content"):
            content = f.get_content()
        else:
            import os
            local_path = frappe.get_site_path(file_url.lstrip("/"))
            with open(local_path, "rb") as fh:
                content = fh.read()
    except Exception as exc:
        parse_log.append({"log_type": "Error", "message": f"Failed to read file: {exc}", "row_index": 0})
        return rows, parse_log

    if not content:
        parse_log.append({"log_type": "Error", "message": "Empty file content", "row_index": 0})
        return rows, parse_log

    # Try XLSX first (more common format)
    if _OPENPYXL:
        try:
            wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            
            # Get headers from first row
            first_row = next(ws.iter_rows(values_only=True))
            headers = [str(x).strip() if x is not None else "" for x in first_row]
            
            # Find column indices
            date_idx = _get_header_index(headers, "statement_date", "date", "statement date")
            ref_idx = _get_header_index(headers, "customer_reference", "reference", "customer reference")
            debit_idx = _get_header_index(headers, "file_debit_amount", "debit", "file debit amount")
            credit_idx = _get_header_index(headers, "file_credit_amount", "credit", "file credit amount")
            company_idx = _get_header_index(headers, "company")
            currency_idx = _get_header_index(headers, "currency")
            
            if date_idx is None:
                parse_log.append({"log_type": "Error", "message": "Date column not found. Expected 'statement_date' or 'Date'", "row_index": 0})
                return rows, parse_log
            
            # Process data rows
            row_num = 2  # Row 2 in Excel (after header)
            for row_values in ws.iter_rows(values_only=True, min_row=2):
                # Skip empty rows
                if all([c is None or str(c).strip() == "" for c in row_values]):
                    continue
                
                try:
                    # Extract date
                    date_val = _normalize_date(row_values[date_idx] if date_idx < len(row_values) else None)
                    if not date_val:
                        raise ValueError(f"Invalid date value: {row_values[date_idx] if date_idx < len(row_values) else 'None'}")
                    
                    # Extract other fields
                    ref_val = str(row_values[ref_idx]).strip() if ref_idx is not None and ref_idx < len(row_values) and row_values[ref_idx] is not None else ""
                    debit_val = flt(row_values[debit_idx]) if debit_idx is not None and debit_idx < len(row_values) else 0.0
                    credit_val = flt(row_values[credit_idx]) if credit_idx is not None and credit_idx < len(row_values) else 0.0
                    company_val = str(row_values[company_idx]).strip() if company_idx is not None and company_idx < len(row_values) and row_values[company_idx] is not None else None
                    currency_val = str(row_values[currency_idx]).strip() if currency_idx is not None and currency_idx < len(row_values) and row_values[currency_idx] is not None else None
                    
                    rows.append({
                        "statement_date": date_val,
                        "customer_reference": ref_val,
                        "file_debit_amount": debit_val,
                        "file_credit_amount": credit_val,
                        "company": company_val,
                        "currency": currency_val,
                        "row_index": row_num
                    })
                except Exception as e:
                    parse_log.append({
                        "log_type": "Error",
                        "message": f"Row {row_num}: {str(e)}",
                        "row_index": row_num
                    })
                
                row_num += 1
            
            return rows, parse_log
            
        except Exception as exc:
            parse_log.append({"log_type": "Error", "message": f"Failed to parse XLSX: {exc}", "row_index": 0})
            # Don't return yet, try CSV as fallback

    # CSV attempt
    try:
        text = content.decode("utf-8-sig") if isinstance(content, bytes) else str(content)
        reader = csv.DictReader(io.StringIO(text))
        if reader.fieldnames:
            headers = [h.strip() if h else "" for h in reader.fieldnames]
            
            for i, r in enumerate(reader, start=2):  # Start at 2 (row 1 is header)
                try:
                    # Try multiple header variations
                    date_raw = r.get("statement_date") or r.get("Date") or r.get("date") or r.get("statement date")
                    date_val = _normalize_date(date_raw)
                    if not date_val:
                        raise ValueError(f"Invalid date value: {date_raw}")
                    
                    rows.append({
                        "statement_date": date_val,
                        "customer_reference": (r.get("customer_reference") or r.get("Reference") or r.get("reference") or "").strip(),
                        "file_debit_amount": flt(r.get("file_debit_amount") or r.get("Debit") or r.get("debit") or 0.0),
                        "file_credit_amount": flt(r.get("file_credit_amount") or r.get("Credit") or r.get("credit") or 0.0),
                        "company": (r.get("company") or r.get("Company") or "").strip() or None,
                        "currency": (r.get("currency") or r.get("Currency") or "").strip() or None,
                        "row_index": i
                    })
                except Exception as e:
                    parse_log.append({
                        "log_type": "Error",
                        "message": f"Row {i}: {str(e)}",
                        "row_index": i
                    })
            return rows, parse_log
    except Exception as exc:
        parse_log.append({"log_type": "Error", "message": f"Failed to parse CSV: {exc}", "row_index": 0})

    if not rows and not parse_log:
        parse_log.append({"log_type": "Error", "message": "Unsupported file format or empty file.", "row_index": 0})
    
    return rows, parse_log


def _build_preview(doc, rows, parse_log=None):
    parse_log = parse_log or []
    doc.set("debtor_import_preview", [])
    doc.set("import_log", [])

    for log in parse_log:
        doc.append("import_log", log)

    seen = set()
    for r in rows:
        # Date should already be normalized from _parse_file
        statement_date = r.get("statement_date")
        if not statement_date or not isinstance(statement_date, date):
            doc.append("import_log", {
                "log_type": "Error",
                "message": f"Row {r.get('row_index')}: Invalid date object",
                "row_index": r.get("row_index")
            })
            continue

        debit = flt(r.get("file_debit_amount"))
        credit = flt(r.get("file_credit_amount"))

        unique_hash = _hash_row(r.get("company"), r.get("customer_reference"),
                                statement_date, credit, debit)
        is_duplicate = 1 if unique_hash in seen else 0
        seen.add(unique_hash)

        doc.append("debtor_import_preview", {
            "statement_date": statement_date,
            "customer_reference": r.get("customer_reference"),
            "file_debit_amount": debit,
            "file_credit_amount": credit,
            "target_payment": 0.0,
            "target_invoice": 0.0,
            "row_index": r.get("row_index"),
            "unique_hash": unique_hash,
            "is_duplicate": is_duplicate,
            "company": r.get("company"),
            "currency": r.get("currency")
        })

        if is_duplicate:
            doc.append("import_log", {
                "log_type": "Warning",
                "message": _("Duplicate row in file"),
                "row_index": r.get("row_index")
            })

    doc.parsed_rows = len(doc.debtor_import_preview)
    doc.save()


def _create_statement_entries(doc):
    created = 0
    skipped = 0
    for preview in doc.debtor_import_preview:
        if preview.is_duplicate:
            skipped += 1
            continue
        if frappe.db.exists("Debtor Statement Entry", {"unique_hash": preview.unique_hash}):
            skipped += 1
            doc.append("import_log", {
                "log_type": "Info",
                "message": _("Skipped existing statement entry"),
                "row_index": preview.row_index
            })
            continue
        d = frappe.new_doc("Debtor Statement Entry")
        d.company = preview.company or doc.get("company") or frappe.db.get_default("company")
        d.customer = doc.get("customer")
        d.statement_date = preview.statement_date
        d.payment_amount_credit = flt(preview.file_credit_amount)
        d.payment_amount_debit = flt(preview.file_debit_amount)
        d.customer_reference = preview.customer_reference
        d.currency = preview.currency or frappe.db.get_default("currency")
        d.source_import = doc.name
        d.unique_hash = preview.unique_hash
        d.is_reconciled = 0
        d.insert(ignore_permissions=True)
        created += 1

    doc.save()
    return created, skipped